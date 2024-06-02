import openpyxl
import os
class HomePageData:
    @staticmethod
    def getTestData(Test_Case_Name):
        book = openpyxl.load_workbook(os.curdir + "\\testData\\" + "PythonDemoData.xlsx")
        sheet = book.active
        Dict = {}
        for r in range(1, sheet.max_row + 1):
            if sheet.cell(row=r, column=1).value == Test_Case_Name:
                for c in range(2, sheet.max_column + 1):
                    # Dict["Name"] = "Nazma"
                    Dict[sheet.cell(row=1, column=c).value] = sheet.cell(row=r, column=c).value
        return [Dict]  # As the test data is expected in list format with =in list the data wrapped as
        # multiple data sets in dictionary form

